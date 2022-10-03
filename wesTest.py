import openpyxl
from PyPDF2 import PdfFileReader
import re
import courseParser

filePath = "Sample_Input5.pdf"

pdf = PdfFileReader(filePath)
useableLines = []
for pageNum in range(pdf.numPages):
    pageObj = pdf.getPage(pageNum)
    #print("Page Number: ",pageNum)
    txt = pageObj.extractText().split("\n") #Split PDF String by \n
    for i in range(len(txt)):
        #print (txt[i])
        if( "Still Needed:" in txt[i]):#If line says 'Still Needed:', add next line to list
            useableLines.append(txt[i+1])
#print(useableLines)



Dictionary = {}
for x in range(len(useableLines)):
    words = useableLines[x].split()#Split each line word by word
    if "PEDS" in words:#if PEDS Course is needed, create course object
        course= courseParser.Course("PEDS 2378",1,False)
        Dictionary[course.CourseNum] = course
    elif "LEAD" in words:#if LEAD course is needed, create Course object
        course= courseParser.Course("LEAD 1705",2,False)
        Dictionary[course.CourseNum] = course
    elif (words[-1] == "or"):#If credit can be satisfied by another class, move to next class
        pass
    else:
        result = re.search(r"\d{4}K?(?!.*\d{4}K?)", useableLines[x])#Search for 4 digits in a row. eg:3108 or 1301K
        result2 = re.search('[A-Z]+[A-Z]+[A-Z]+[A-Z](?!.*[A-Z]+[A-Z]+[A-Z]+[A-Z])', useableLines[x])# search for 4 capital letters, eg:CPSC or CYBR
        result3= re.search(r"\d{1}@", useableLines[x])#search for first single digit
        if (result and result2):#if first 2 are met, create course accordingly, eg: CYBR 3108
            code = result.group()
            title = result2.group()
            num= title + " " + code
            cr = re.match("Credits", useableLines[x])
            if (cr):#if a certain amount of credits are required, create the course with that many hours. eg: 4 credits in BIOL 1225K
                hours = re.search(r'\d{1}', useableLines[x])
                
                course = courseParser.Course(num,hours,False)
            elif(code.endswith("K")):#if course ends with K, it is a lab and has 4 credit hours
                course = courseParser.Course(num, 4, False)
            else:#Otherwise, class is 3 credit hours
                course = courseParser.Course(num, 3, False)
            Dictionary[num] = course #Add course to Dictionary
            #print("Matching Word:",result2.group(), result.group() + " : Credits = " + str(cr))
            
        elif (result2 and result3):#if lines has 4 capital letters and digit@ it is an elective
            title= result2.group()
            code = " Elective"
            hours = int(re.search(r'\d{1}', useableLines[x]).group())
            #print(code+ " " +str(result2.group()) +  " "+ str(hours))
            #print("before dividing result2: " + str(hours))
            hours = int(hours /3)
            #print("after dividing result2: "+ str(hours))
            for x in range(hours):#Divide number of credits needed by 3 to get number of elective classes are needed
                num= title + code+ str(x+1)
                course = courseParser.Course(num, 3, False)
                Dictionary[course.CourseNum] = course#add course to Dictionary
        elif (result3):#If only a digit@ is found, it is a general elective
            code= "Elective"
            hours = int(re.search(r'\d{1}', useableLines[x]).group())
            #print("before dividing result3: " + str(hours))
            hours = int(hours / 3)
            #print("after dividing result3: " + str(hours))
            for x in range(hours):#Divide number of credits needed by 3 to get number of elective classes are needed
                code= code + str(x + 1)
                course = courseParser.Course(code, 3, False)
                Dictionary[course.CourseNum] = course #Add course to Dictionary
#print(title + " " + str(code) + " " + str(hours))
myDict = Dictionary.keys()
#print("x = " +str(x))


wb= openpyxl.load_workbook("CPSCXL.xlsx")
sheet= wb.active

col_max= sheet.max_column
row_max= sheet.max_row
Dictionary = {}
for i in range (1,row_max + 1): #Read 1st 2 columns of xl file and create courses accordingly
    CourseNum= sheet.cell(i, 1).value
    Hours= sheet.cell(i,2).value
    if CourseNum in myDict:
        course= courseParser.Course(CourseNum,Hours,False)#If the course is already in the courses parsed from degreeworks, the class is created as not being completed
        #print(CourseNum+ " "+ str(Hours))
    else:
        course= courseParser.Course(CourseNum,Hours,True)#If the course is not already in the courses parsed from degreeworks, the class is created as being completed
        #print(CourseNum+ " "+ str(Hours))
    Dictionary[course.CourseNum] = course
    print(course.CourseNum+ " "+ str(Hours))
    print(i)

for i in range(1, row_max + 1):
    CourseNum= sheet.cell(i, 1).value
    for y in range(3, col_max+1):
        pre_req = sheet.cell(i, y).value
        if pre_req !="none":
            try:#if the key exists
                Dictionary[CourseNum].prereq.append(Dictionary[pre_req])#reiterate adding newly created course objects into a list for each of its pre-requisites.
            except:
                pass
    #myDict[CourseNum] = Dictionary[CourseNum]#put all classes from the Excel file into the dictionary of needed courses
    #print(CourseNum+ " "+ str(Hours))
#print(parsedDict["CPSC 5128"].CourseNum +" "+ str(parsedDict["CPSC 5128"].CreditHours) + " Completed: "+ str(parsedDict["CPSC 5128"].isComplete()))#Debug tool
for each in myDict:
    print(each)
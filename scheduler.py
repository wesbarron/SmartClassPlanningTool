import networkx as nx
import re
import courseParser
import worksheet
import GUI
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Color, PatternFill, Font, Border
from openpyxl.styles import colors
from openpyxl.cell import Cell

def getTotalCreditHours(creditHours, CoursesDict):
    coursesInSchedule = []
    coursesChecklist = []
    Survey = CoursesDict["CPSC 4000"]
    if Survey:
        print("SEMESTER COUNT RETRIEVED")
    for course in CoursesDict: 
        if course != Survey.CourseNum:
            coursesChecklist.append(CoursesDict[course])

    schedule = []
 
    #while(len(coursesChecklist)>0):
    semesterCreditHours = 0
    newSemester = []
    poppedCourseIndexes = []
    for i in range(0, len(coursesChecklist)):
        #if credit hours limit is hit
        course = coursesChecklist[i]
        
        if(course.Completed == False):
            if((semesterCreditHours + course.CreditHours)<=creditHours):
                
                if course.canBeTaken():
                    #print("Course is being appended: " + course.CourseNum)
                    poppedCourseIndexes.append(i)
                    newSemester.append(course)
                    semesterCreditHours += course.CreditHours
                    coursesInSchedule.append(course)
    #print (semesterCreditHours)
    return semesterCreditHours

def setSchedule(creditHours, CoursesDict):
    print("Setting Schedule: ")
    coursesInSchedule = []
    coursesChecklist = []
    coursesInWorkBook = []

    for course in CoursesDict: 
       
        coursesChecklist.append(CoursesDict[course])

    schedule = []

    #print schedule

    wb = load_workbook('schedule.xlsx')
    #wb = Workbook()
    ws = wb.active
    ws.title = "Path To Graduation"
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    ws.sheet_view.showGridLines = True
    # redFill = PatternFill(start_color='FFFF0000',
    #                end_color='FFFF0000',
    #                fill_type='solid')
    hoursCount = GUI.globalHoursCount
    print(getTotalCreditHours(hoursCount, CoursesDict))
    creditHoursAllowed = getTotalCreditHours(hoursCount, CoursesDict)

    
    while(len(coursesChecklist)>0):
        semesterCreditHours = 0
        newSemester = []
        poppedCourseIndexes = []
        workBookHours = 0
        logHours = 0
        secondLogHours = 0
        thirdLogHours = 0
        fourthLogHours = 0
        fifthLogHours = 0
        sixthLogHours = 0
    
        for i in range(0, len(coursesChecklist)):
            #if credit hours limit is hit
            course = coursesChecklist[i]
            
                    
            if(course.Completed == False):
                if((semesterCreditHours + course.CreditHours)<=creditHours):
                    for prereqCourse in course.prereq:
                        for courseInSchedule in coursesInSchedule:
                            if(prereqCourse == courseInSchedule): #if prereq course is in schedule already
                                #if prereq course is in this semester, it cannot be taken
                                exists = courseInSchedule in newSemester
                                if(exists == False): #else it is in previous semester, it can be taken
                                    prereqCourse.Completed = True
                    print(course.canBeTaken())
                    if course.canBeTaken():
                        print("Course is being appended: " + course.CourseNum)
                        poppedCourseIndexes.append(i)
                        newSemester.append(course)
                        semesterCreditHours += course.CreditHours
                        coursesInSchedule.append(course)
                        coursesInWorkBook.clear()

                        
                
            else:
                poppedCourseIndexes.append(i)
            poppedCourseIndexes.sort(reverse=True)
        if(len(newSemester)!=0):
            for index in poppedCourseIndexes:
                coursesChecklist.pop(index)
            schedule.append(newSemester)


    # if(len(coursesChecklist)>0):
    #     for course in coursesChecklist:
    #         if((semesterCreditHours+course.CreditHours)>creditHours):
    #             schedule.append(newSemester)
    #             break
    #         #Check for courses with prereqs if prereqs are in coursesInSchedule array
    #         if(len(course.prereq) > 0):
    #             for prereqCourse in course.prereq:
    #                 for courseInSchedule in coursesInSchedule:
    #                     if(prereqCourse == courseInSchedule):
    #                         exists = courseInSchedule in newSemester
    #                         if(exists == False):
    #                             #takes prereq off the prereq array
    #                             course.prereq.remove(prereqCourse)
    #         #Check for courses without prereq
    #         elif(len(course.prereq)==0):
    #             courseFromChecklist = coursesChecklist.pop(0)
    #             newSemester.append(courseFromChecklist)
    #             coursesInSchedule.append(courseFromChecklist)
    #             semesterCreditHours += course.CreditHours 

    # totalColumn = len(coursesInSchedule)
    # ws.cell(totalColumn+3,1).value = "Total"
    # ws.cell(totalColumn+3,2).value = creditHoursAllowed
    # ws.cell(totalColumn+3,1).fill = redFill
    # ws.cell(totalColumn+3,2).fill = redFill
    # ws.cell(totalColumn+3,1).font = Font(color=colors.WHITE)
    # ws.cell(totalColumn+3,2).font = Font(color=colors.WHITE)
    # worksheet.set_border(ws, "A1:B"+str(totalColumn+3))

    # ws.cell(totalColumn+3,4).value = "Total"
    # ws.cell(totalColumn+3,5).value = creditHoursAllowed
    # ws.cell(totalColumn+3,4).fill = redFill
    # ws.cell(totalColumn+3,5).fill = redFill
    # ws.cell(totalColumn+3,4).font = Font(color=colors.WHITE)
    # ws.cell(totalColumn+3,5).font = Font(color=colors.WHITE)
    # worksheet.set_border(ws, "D1:E"+str(totalColumn+3))
    first = 1
    second = 1
    third = 1
    fourth = 1
    fifth = 1
    sixth = 1    
    print(schedule)
    for semester in schedule:
        print("New Semester")
        for course in semester:
            coursePrereqStr = ""
            for prereqCourse in course.prereq:
                coursePrereqStr += prereqCourse.CourseNum

            print(course.CourseNum + " : " + str(course.CreditHours) + " : " + coursePrereqStr)

            ## Results saved to workbook ##

            if ((logHours) <= creditHoursAllowed):
                if coursePrereqStr != "" and prereqCourse.CourseNum not in coursesInSchedule:
                    
                    ws.cell(first+1,1).value = coursePrereqStr
                    ws.cell(first+1,2).value = course.CreditHours
                    #print("This is the course number : " + course.CourseNum)
                    workBookHours += course.CreditHours
                    coursesInWorkBook.append(course.CourseNum)
                else:
                    
                    ws.cell(first+1,1).value = course.CourseNum
                    ws.cell(first+1,2).value = course.CreditHours
                    #print("This is the course number : " + course.CourseNum)
                    workBookHours += course.CreditHours
                    coursesInWorkBook.append(course.CourseNum) 
                logHours = workBookHours + course.CreditHours
                print("log Hours first if: " + str(logHours))
                print("courses in workBook 1st time: " + str(coursesInWorkBook))
                first = first+1

            if ((secondLogHours) <= (creditHoursAllowed*2) and course.CourseNum not in coursesInWorkBook and course.CourseNum != "CPSC 4000"):
                
                # if coursePrereqStr != "" and prereqCourse.CourseNum not in coursesInSchedule:
                    
                #     ws.cell(second+1,3).value = coursePrereqStr
                #     ws.cell(second+1,4).value = course.CreditHours
                #     #print("This is the course number : " + course.CourseNum)
                #     workBookHours += course.CreditHours
                #     coursesInWorkBook.append(course.CourseNum)
                # else:
                    
                ws.cell(second+1,3).value = course.CourseNum
                ws.cell(second+1,4).value = course.CreditHours
                #print("This is the course number : " + course.CourseNum)
                workBookHours += course.CreditHours
                coursesInWorkBook.append(course.CourseNum) 
                secondLogHours = workBookHours + course.CreditHours
                print("log Hours second if: " + str(secondLogHours))
                print("courses in workBook 2nd time: " + str(coursesInWorkBook))
                #print("log hours + courseHours : " + str(logHours+course.CreditHours))
                second = second+1

            if ((thirdLogHours) <= (creditHoursAllowed*3) and course.CourseNum not in coursesInWorkBook and thirdLogHours <= 45):
                # if coursePrereqStr != "" and prereqCourse.CourseNum not in coursesInSchedule:
                    
                #     ws.cell(third+1,5).value = coursePrereqStr
                #     ws.cell(third+1,6).value = course.CreditHours
                #     #print("This is the course number : " + course.CourseNum)
                #     workBookHours += course.CreditHours
                #     coursesInWorkBook.append(course.CourseNum)
                # else:
                    
                ws.cell(third+1,5).value = course.CourseNum
                ws.cell(third+1,6).value = course.CreditHours
                #print("This is the course number : " + course.CourseNum)
                workBookHours += course.CreditHours
                coursesInWorkBook.append(course.CourseNum) 
                thirdLogHours = workBookHours + course.CreditHours
                print("log Hours third if: " + str(thirdLogHours))
                print("courses in workBook 3rd time: " + str(coursesInWorkBook))
                #print("log hours + courseHours : " + str(logHours+course.CreditHours))
                third = third+1

            if ((fourthLogHours) <= (creditHoursAllowed*4) and course.CourseNum not in coursesInWorkBook):
                if coursePrereqStr != "" and prereqCourse.CourseNum not in coursesInSchedule:
                    
                    ws.cell(fourth+1,5).value = coursePrereqStr
                    ws.cell(fourth+1,6).value = course.CreditHours
                    #print("This is the course number : " + course.CourseNum)
                    workBookHours += course.CreditHours
                    coursesInWorkBook.append(course.CourseNum)
                else:
                    
                    ws.cell(fourth+1,5).value = course.CourseNum
                    ws.cell(fourth+1,6).value = course.CreditHours
                    #print("This is the course number : " + course.CourseNum)
                    workBookHours += course.CreditHours
                    coursesInWorkBook.append(course.CourseNum) 
                fourthLogHours = workBookHours + course.CreditHours
                print("log Hours third if: " + str(fourthLogHours))
                print("courses in workBook 3rd time: " + str(coursesInWorkBook))
                #print("log hours + courseHours : " + str(logHours+course.CreditHours))
                fourth = fourth+1

            if ((fifthLogHours) <= (creditHoursAllowed*5) and course.CourseNum not in coursesInWorkBook):
                if coursePrereqStr != "" and prereqCourse.CourseNum not in coursesInSchedule:
                    
                    ws.cell(fifth+1,5).value = coursePrereqStr
                    ws.cell(fifth+1,6).value = course.CreditHours
                    #print("This is the course number : " + course.CourseNum)
                    workBookHours += course.CreditHours
                    coursesInWorkBook.append(course.CourseNum)
                else:
                    
                    ws.cell(fifth+1,5).value = course.CourseNum
                    ws.cell(fifth+1,6).value = course.CreditHours
                    #print("This is the course number : " + course.CourseNum)
                    workBookHours += course.CreditHours
                    coursesInWorkBook.append(course.CourseNum) 
                fifthLogHours = workBookHours + course.CreditHours
                print("log Hours third if: " + str(fifthLogHours))
                print("courses in workBook 3rd time: " + str(coursesInWorkBook))
                #print("log hours + courseHours : " + str(logHours+course.CreditHours))
                fifth = fifth+1

            if ((sixthLogHours) <= (creditHoursAllowed*6) and course.CourseNum not in coursesInWorkBook):
                if coursePrereqStr != "" and prereqCourse.CourseNum not in coursesInSchedule:
                    
                    ws.cell(sixth+1,5).value = coursePrereqStr
                    ws.cell(sixth+1,6).value = course.CreditHours
                    #print("This is the course number : " + course.CourseNum)
                    workBookHours += course.CreditHours
                    coursesInWorkBook.append(course.CourseNum)
                else:
                    
                    ws.cell(sixth+1,5).value = course.CourseNum
                    ws.cell(sixth+1,6).value = course.CreditHours
                    #print("This is the course number : " + course.CourseNum)
                    workBookHours += course.CreditHours
                    coursesInWorkBook.append(course.CourseNum) 
                sixthLogHours = workBookHours + course.CreditHours
                print("log Hours third if: " + str(sixthLogHours))
                print("courses in workBook 3rd time: " + str(coursesInWorkBook))
                #print("log hours + courseHours : " + str(logHours+course.CreditHours))
                sixth = sixth+1

            wb.save("scheduleResult.xlsx")
                        
                
        

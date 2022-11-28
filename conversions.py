import courseParser
from openpyxl import Workbook, load_workbook

def replaceElectives(schedule, caseSchedule):
    caseElectives = []
    caseElectivesIndex = 0
    for semester in caseSchedule:
       for course in semester:
        if course[-1] == 'E' and course[-2] != 'G':
            caseElectives.append(course)
    
    for semester in schedule:
        for x in range(0,len(semester)):
            if "Elective" in semester[x].CourseNum and ('CYBR' in semester[x].CourseNum or 'CPSC' in semester[x].CourseNum):
                semester[x] = caseElectives[caseElectivesIndex][:-1] + " Elective"
                if not (caseElectivesIndex + 1 > len(caseElectives)):
                    caseElectivesIndex += 1
    return schedule
def printSchedule(schedule):
    for semester in schedule:
        for course in semester:
            if(isinstance(course,str)):
                print(course)
            elif (isinstance(course, courseParser.Course)):
                print(course.CourseNum)
            
def buildSchedule(schedule, hours, fileName):

    wb = load_workbook('schedule.xlsx')
    ws = wb.active
    ws.title = "Path To Graduation"
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 20
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 25
    ws.column_dimensions['F'].width = 20
    ws.sheet_view.showGridLines = True

    first = 1
    second = 1
    third = 1
    fourth = 10
    fifth = 10
    sixth = 10 
    count = 0
    total1 = 0
    total2 = 0
    total3 = 0
    total4 = 0
    total5 = 0
    total6 = 0

    coursesInWorkBook = []
    totalHoursCounter = []
    electives = []

    for semester in schedule:
        for course in semester:
            
            if (isinstance(course, courseParser.Course)): # Insert all regualar courses into excel (not "CPSC 4000", saving for graduating semester)
            
                count += course.CreditHours

                if ((count) <= hours) and course.CourseNum not in coursesInWorkBook and course.CourseNum != "CPSC 4000":
                    ws.cell(first+1,1).value = course.CourseNum
                    ws.cell(first+1,2).value = course.CreditHours
                    total1 += ws.cell(first+1,2).value
                    first = first+1
                    coursesInWorkBook.append(course.CourseNum)
                    totalHoursCounter.append(course.CreditHours)
                    
                

                elif (count) <= (hours*2) and course.CourseNum not in coursesInWorkBook and course.CourseNum != "CPSC 4000":
                    ws.cell(second+1,3).value = course.CourseNum
                    ws.cell(second+1,4).value = course.CreditHours
                    total2 += ws.cell(second+1,4).value
                    second = second+1
                    coursesInWorkBook.append(course.CourseNum)
                    totalHoursCounter.append(course.CreditHours)
                    
                
                    
                elif (count) <= (hours*3) and course.CourseNum not in coursesInWorkBook and course.CourseNum != "CPSC 4000":
                    ws.cell(third+1,5).value = course.CourseNum
                    ws.cell(third+1,6).value = course.CreditHours
                    total3 += ws.cell(third+1,6).value
                    third = third+1
                    coursesInWorkBook.append(course.CourseNum)
                    totalHoursCounter.append(course.CreditHours)
                    

                elif (count) <= (hours*4) and course.CourseNum not in coursesInWorkBook and course.CourseNum != "CPSC 4000":
                    ws.cell(fourth+1,1).value = course.CourseNum
                    ws.cell(fourth+1,2).value = course.CreditHours
                    total4 += ws.cell(fourth+1,2).value
                    fourth = fourth+1
                    coursesInWorkBook.append(course.CourseNum)
                    totalHoursCounter.append(course.CreditHours)
                    


                elif (count) <= (hours*5) and course.CourseNum not in coursesInWorkBook and course.CourseNum != "CPSC 4000":
                    ws.cell(fifth+1,3).value = course.CourseNum
                    ws.cell(fifth+1,4).value = course.CreditHours
                    total5 += ws.cell(fifth+1,4).value
                    fifth = fifth+1
                    coursesInWorkBook.append(course.CourseNum)
                    totalHoursCounter.append(course.CreditHours)
                    


                elif (count) <= (hours*6) and course.CourseNum not in coursesInWorkBook and course.CourseNum != "CPSC 4000":
                    ws.cell(sixth+1,5).value = course.CourseNum
                    ws.cell(sixth+1,6).value = course.CreditHours
                    total6 += ws.cell(sixth+1,6).value
                    sixth = sixth+1
                    coursesInWorkBook.append(course.CourseNum)
                    totalHoursCounter.append(course.CreditHours)

                #counterSum = sum(totalHoursCounter) # leaving for debugging
        
                
            elif (isinstance(course,str)):
                electives.append(course)

                
    for i in electives: # If the course is an elective and there is enough eligible hours left in the semester
        if (total1 + total2 < hours) and i not in coursesInWorkBook:
            ws.cell(first+1,1).value = i
            ws.cell(first+1,2).value = 3
            first = first+1
            coursesInWorkBook.append(i)
            total1 += 3 
        elif (total2 + total3 < hours) and i not in coursesInWorkBook:
            ws.cell(second+1,3).value = i
            ws.cell(second+1,4).value = 3
            second = second+1
            coursesInWorkBook.append(i)
            total2 += 3
        elif (total3 + total4 < hours) and i not in coursesInWorkBook:
            ws.cell(third+1,5).value = i
            ws.cell(third+1,6).value = 3
            third = third+1
            coursesInWorkBook.append(i)
            total3 += 3
        elif (total4 + total5 < hours) and i not in coursesInWorkBook:
            ws.cell(fourth+1,1).value = i
            ws.cell(fourth+1,2).value = 3
            fourth = fourth+1
            coursesInWorkBook.append(i)
            total4 += 3
        elif (total5 + total6 < hours) and i not in coursesInWorkBook:
            ws.cell(fifth+1,3).value = i
            ws.cell(fifth+1,4).value = 3
            fifth = fifth+1
            coursesInWorkBook.append(i)
            total5 += 3
        elif (total6 + 3 < hours) and i not in coursesInWorkBook:
            ws.cell(sixth+1,3).value = i
            ws.cell(sixth+1,4).value = 3
            sixth = sixth+1
            coursesInWorkBook.append(i)
            total6 += 3             

    # Now find the graduating semester and insert "CPSC 4000"
    if (total1 + total2) < (hours) and "CPSC 4000" not in coursesInWorkBook:
        ws.cell(first+1,1).value = "CPSC 4000"
        ws.cell(first+1,2).value = 0
        first = first+1
        coursesInWorkBook.append("CPSC 4000") 

    elif (total2 + total3) < (hours) and "CPSC 4000" not in coursesInWorkBook:
        ws.cell(second+1,3).value = "CPSC 4000"
        ws.cell(second+1,4).value = 0
        second = second+1
        coursesInWorkBook.append("CPSC 4000") 

    elif (total3 + total4) < (hours) and "CPSC 4000" not in coursesInWorkBook:
        ws.cell(third+1,5).value = "CPSC 4000"
        ws.cell(third+1,6).value = 0
        third = third+1
        coursesInWorkBook.append("CPSC 4000")

    elif (total4 + total5) < (hours) and "CPSC 4000" not in coursesInWorkBook:
        ws.cell(fourth+1,1).value = "CPSC 4000"
        ws.cell(fourth+1,2).value = 0
        fourth = fourth+1
        coursesInWorkBook.append("CPSC 4000")

    elif (total5 + total6) < (hours) and "CPSC 4000" not in coursesInWorkBook:
        ws.cell(fifth+1,3).value = "CPSC 4000"
        ws.cell(fifth+1,4).value = 0
        fifth = fifth+1
        coursesInWorkBook.append("CPSC 4000")

    elif (total6) < (hours) and "CPSC 4000" not in coursesInWorkBook:
        ws.cell(sixth+1,3).value = "CPSC 4000"
        ws.cell(sixth+1,4).value = 0
        sixth = sixth+1
        coursesInWorkBook.append("CPSC 4000")


    wb.save(fileName)
    print("Please open the files " + fileName + " to view your schedule.")

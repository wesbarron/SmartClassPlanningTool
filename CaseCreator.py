from openpyxl import load_workbook


def createCaseSchedules(CaseLibrary): 
    wb = load_workbook(CaseLibrary)

    ws = wb.active
    Web_Track = []
    Games_Track = []
    CyberSecurity_Track = []
    Software_Track = []

    for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row, values_only=False):
        #print ("NEW ROW\n")DEBUG LINE
        Schedule_case = []
        for cell in row:
            if cell.column ==1:
                Track = cell.value
            else:
                String = (cell.value)

                if String != None:
                    Semester = String.split(", ")
                    if len(Semester) > 0:
                        Schedule_case.append(Semester)
        match Track:
            case "Games":
                Games_Track.append(Schedule_case)
                continue
            case "Web":
                Web_Track.append(Schedule_case)
                continue
            case "Software":
                Software_Track.append(Schedule_case)
                continue
            case "CyberSecurity":
                CyberSecurity_Track.append(Schedule_case)
                continue
    caseScheduleTracks = {"WebTrack" : Web_Track, "GamesTrack" : Games_Track, "CyberSecurityTrack" : CyberSecurity_Track, "SoftwareTrack" : Software_Track}
    return caseScheduleTracks
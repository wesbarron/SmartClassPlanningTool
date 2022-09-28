import openpyxl
from PyPDF2 import PdfFileReader
import re



def getContent(filePath):#Parses Degree Works
    pdf = PdfFileReader(filePath)
    useableLines = []
    for pageNum in range(pdf.numPages):
        pageObj = pdf.getPage(pageNum)
        #print("Page Number: ",pageNum)
        txt = pageObj.extractText().split("\n") #Split PDF String by \n
        for i in range(len(txt)):
            #print (txt[i])
            if( "Still Needed:" in txt[i]):#If line says 'Still Needed:', add next line to list
                useableLines.append(txt[i+1])
    #print(useableLines)

    return useableLines# Return all lines with needed Course Codes


def createFromParse(lines):
    Dictionary = {}
    for x in range(len(lines)):
        words = lines[x].split()#Split each line word by word
        if "PEDS" in words:#if PEDS Course is needed, create course object
            course= Course("PEDS 2378",1,False)
            Dictionary[course.CourseNum] = course
        elif "LEAD" in words:#if LEAD course is needed, create Course object
            course= Course("LEAD 1705",2,False)
            Dictionary[course.CourseNum] = course
        elif (words[-1] == "or"):#If credit can be satisfied by another class, move to next class
            pass
        else:
            result = re.search(r"\d{4}K?(?!.*\d{4}K?)", lines[x])#Search for 4 digits in a row. eg:3108 or 1301K
            result2 = re.search('[A-Z]+[A-Z]+[A-Z]+[A-Z](?!.*[A-Z]+[A-Z]+[A-Z]+[A-Z])', lines[x])# search for 4 capital letters, eg:CPSC or CYBR
            result3= re.search(r"\d{1}@", lines[x])#search for first single digit
            if (result and result2):#if first 2 are met, create course accordingly, eg: CYBR 3108
                code = result.group()
                title = result2.group()
                num= title + " " + code
                cr = re.match("Credits", lines[x])
                if (cr):#if a certain amount of credits are required, create the course with that many hours. eg: 4 credits in BIOL 1225K
                    hours = re.search(r'\d{1}', lines[x])
                    course = Course(num,hours,False)
                elif(code.endswith("K")):#if course ends with K, it is a lab and has 4 credit hours
                    course = Course(num, 4, False)
                else:#Otherwise, class is 3 credit hours
                    course = Course(num, 3, False)
                Dictionary[num] = course #Add course to Dictionary
                #print("Matching Word:",result2.group(), result.group())
            elif (result2 and result3):#if lines has 4 capital letters and digit@ it is an elective
                title= result2.group()
                code = " Elective"
                hours = int(re.search(r'\d{1}', lines[x]).group())
                hours = int(hours /3)
                for x in range(hours):#Divide number of credits needed by 3 to get number of elective classes are needed
                    num= title + code+ str(x+1)
                    course = Course(num, 3, False)
                    Dictionary[course.CourseNum] = course#add course to Dictionary
            elif (result3):#If only a digit@ is found, it is a general elective
                code= "Elective"
                hours = int(re.search(r'\d{1}', lines[x]).group())
                hours = int(hours / 3)
                for x in range(hours):#Divide number of credits needed by 3 to get number of elective classes are needed
                    code= code + str(x + 1)
                    course = Course(code, 3, False)
                    Dictionary[course.CourseNum] = course #Add course to Dictionary
    return Dictionary


def readXL(filePath,parsedDict):
    wb= openpyxl.load_workbook(filePath)
    sheet= wb.active

    col_max= sheet.max_column
    row_max= sheet.max_row
    Dictionary = {}
    for i in range (1,row_max + 1): #Read 1st 2 columns of xl file and create courses accordingly
        CourseNum= sheet.cell(i, 1).value
        Hours= sheet.cell(i,2).value
        #print(CourseNum+ " "+ str(Hours))
        if CourseNum in parsedDict.keys():
            course= Course(CourseNum,Hours,False)#If the course is already in the courses parsed from degreeworks, the class is created as not being completed
        else:
            course= Course(CourseNum,Hours,True)#If the course is not already in the courses parsed from degreeworks, the class is created as being completed
        Dictionary[course.CourseNum] = course


    for i in range(1, row_max + 1):
        CourseNum= sheet.cell(i, 1).value
        for y in range(3, col_max+1):
            pre_req = sheet.cell(i, y).value
            if pre_req !="none":
                try:#if the key exists
                    Dictionary[CourseNum].prereq.append(Dictionary[pre_req])#reiterate adding newly created course objects into a list for each of its pre-requisites.
                except:
                    pass
        parsedDict[CourseNum] = Dictionary[CourseNum]#put all classes from the Excel file into the dictionary of needed courses
    #print(parsedDict["CPSC 5128"].CourseNum +" "+ str(parsedDict["CPSC 5128"].CreditHours) + " Completed: "+ str(parsedDict["CPSC 5128"].isComplete()))#Debug tool
    return parsedDict

def unfinishedCourses(dictionary):
    for each in dictionary:
        if dictionary[each].Completed == False:
            return False
        else:
            pass
    return True
class Course:
    def __init__(self,CourseNum, CreditHours, completed):
        self.CourseNum = CourseNum
        self.CreditHours = CreditHours
        self.Completed= completed
        self.prereq=[]
        self.takeable= False
    def isComplete(self):
        return self.Completed
    def canBeTaken(self):
        for course in range(len(self.prereq)):
            if(self.prereq[course].isComplete()):
                pass
            else:
                return False
        return True

class Semester:
    def __init__(self,term,hours):
        self.Term=term
        self.MaxHours=hours
        Courses= []
    def scheduledHours(self):
        scheduled= 0
        for x in self.Courses:
            scheduled = scheduled+ self.Courses[x].CreditHour
        return scheduled

    def addClass(self,course):
        if course.canBeTaken and (course.CreditHours + self.scheduledHours() <= self.MaxHours):
            self.Courses.append(course)
            course.Completed = True
            return True
        else:
            return False
    def printSchedule(self):
        for x in self.Courses:
            print (self.Courses[x].CourseNum)

